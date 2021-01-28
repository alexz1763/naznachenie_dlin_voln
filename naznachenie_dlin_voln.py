import openpyxl
import networkx as nx
from numpy import sqrt
import matplotlib.pyplot as plt


# открываю файл с координатами узлов (колонки обозначают: первая название узла, вторая координату по X,
# третья координата по Y)
wb = openpyxl.load_workbook(filename='point.xlsx')
# открываю активный лист
ws = wb.active
# создаю пустой список
data = []
# определяю количество узлов по количеству заполненных строк
N_point = ws.max_row
# собираю значения всех ячеек в список data
for row in ws.values:
   for value in row:
       data.append(value)
# создаю словарь для координат узлов
Nodes = {}
for i in range(0, len(data), 3):
    # заполняю словарь: первое значение ключ - номер узла (с нуля) втрое и третье заначение координаты
    Nodes.setdefault(str(data[i]), (data[i+1], data[i+2]))

# открываю файл с c матрицей инценденций (единица есть ребро, ноль нет ребра, по диагонале считать не нужно)
wb = openpyxl.load_workbook(filename='i-matrix.xlsx')
# открываю активный лист
ws = wb.active
# создаю пустой список
data = []
# проверяю соответствие перечня узлов матрице инценденций
if N_point != ws.max_row:
    print('матрица инценденций не соответствует количеству узлов')
    quit()
N_point = ws.max_row
# собираю значения всех ячеек в список data
for row in ws.values:
   for value in row:
       data.append(value)
# собираю список линий из данных матрицы инценденций
Edges = []
for i in range(0, N_point):
    l = i*5
    for j in range(0, N_point):
        if data[l+j] == 1:
            Edge = (str(i), str(j))
            Edges.append(Edge)

# Создаю граф сети
G = nx.Graph()
# Добавляю в граф сформированный список узлов
G.add_nodes_from(Nodes)
# Добавляю в граф сформированный список линий
G.add_edges_from(Edges)

# расчет длин линий
Lines = []
for edge in G.edges:
    distance = sqrt((Nodes[edge[0]][0] - Nodes[edge[1]][0]) ** 2 + (Nodes[edge[0]][1] - Nodes[edge[1]][1]) ** 2)
    # расчет длин ребер между узлами
    line = (edge[0], edge[1], distance)
    Lines.append(line)
G.add_weighted_edges_from(Lines)
# добавляем длины линий в граф

# собираем направления связи из файла эксел direct.xlsx
# первый столбец номер направления связи, второй столбец исходящий узел, третий входящий узел
wb = openpyxl.load_workbook(filename='direct.xlsx')
# открываю активный лист
ws = wb.active
# создаю пустой список
data = []
# собираю значения всех ячеек в список data
for row in ws.values:
   for value in row:
       data.append(value)
# удаление заголовка таблицы
del data[:3]
# создаю список списков с направлениями связи
Directs = []
for i in range(0, len(data), 3):
    direct = []
    direct.append(str(data[i+1]))
    direct.append(str(data[i+2]))
    Directs.append(direct)

# расчитываю кратчайшие пути для списка направлений связи(библиотека использует алгоритм default = ‘dijkstra’)
shortest_path = []
for direct in Directs:
    shortest_path.append(nx.shortest_path(G, direct[0], direct[1]))

# создаю список линий для каждого кратчашего пути
shortest_path_edge = []
for path in shortest_path:
    shortest_path_edge_lists = []
    for i in range(0, (len(path)-1)):
        shortest_path_edge_lists.append((path[i], path[i+1]))
    shortest_path_edge.append(shortest_path_edge_lists)

# создаю список узлов и линий для графа маршрутов
Path_node = [str(x) for x in range(0, len(shortest_path_edge))]

Path_edges = []
for i in range(0, len(shortest_path_edge)-1):
    for j in range(0, len(shortest_path_edge[i])):
        for k in range(i+1, len(shortest_path_edge)):
            if shortest_path_edge[i][j] in shortest_path_edge[k]:
                Path_edges.append((str(i), str(k)))
            elif shortest_path_edge[i][j][::-1] in shortest_path_edge[k]:
                Path_edges.append((str(i), str(k)))

# создаю граф маршрутов
G_path = nx.Graph()
G_path.add_edges_from(Path_edges)
# раскрашиваю граф
Color_path = nx.coloring.greedy_color(G_path)
# определение максимального количества цветов для раскраски маршрутов
Min_color = int(max(Color_path, key=Color_path.get))+1
# делаем список узлов для каждого цвета
Color_paths_list = []
for i in range(0, Min_color):
    Color_path_list = []
    for key in Color_path:
        if i == Color_path[key]:
            Color_path_list.append(key)
    Color_paths_list.append(Color_path_list)

# найдем независимые неокрашенные маршруты и добавим в список первого цвета
join_path = []
for path in Path_node:
    for color in Color_paths_list:
        if path in color:
            join_path.append(path)
disjoin_path = []
for path in Path_node:
    if path in join_path:
        continue
    else:
        disjoin_path.append(path)
for path in disjoin_path:
    Color_paths_list[0].append(path)

# сохраняем полученный результат в файл exel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'результат'
ws['A1'] = 'Маршруты для направлений связи'
I = 1
for i in range(0, len(shortest_path_edge)):
    I = I + 1
    ws['A'+str(I)] = str(i)
    ws['B'+str(I)] = str(shortest_path_edge[i])
I = I + 1
ws['A' + str(I)] = 'Назначение длин волн для направлений связи'
for i in range(0, len(Color_paths_list)):
    I = I + 1
    ws['A' + str(I)] = str(i+1)
    ws['B' + str(I)] = str(Color_paths_list[i])
I = I + 1
ws['A' + str(I)] = 'Необходимое число длин волн'
I = I + 1
ws['A' + str(I)] = len(Color_paths_list)
wb.save(filename='Results.xlsx')

# выводим графически результат
pos = nx.spring_layout(G)
plt.subplot(121)
plt.title('Исходные данные') # печатаю заголовок рисунка
nx.draw(G, pos, with_labels=True) # рисую граф с названиями и позициями узлов и линий
plt.subplot(122)
plt.title('Распределение длин волн') # печатаю заголовок рисунка
nx.draw(G, pos, with_labels=True) # рисую граф с названиями и позициями узлов и линий
width_line = 15.0 # отрисовываю длины волн
cmaps = ['b', 'r', 'g', 'y', 'o']
for i in range(0, len(Color_paths_list)):
    width_line = width_line - 4.0
    color = cmaps[i]
    for path in Color_paths_list[i]:
        nx.draw_networkx_edges(G, pos, edgelist=shortest_path_edge[int(path)], width=width_line, edge_color=color)# рисую линии на графе
plt.draw() # формирую рисунок
plt.show() # вывожу рисунок


